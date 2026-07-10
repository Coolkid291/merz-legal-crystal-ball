#!/usr/bin/env python3
"""
parse.py - Extract the Merz Legal Crystal Ball workbook into docs/data.json.

Read-only: the .xlsm is opened with data_only=True and is NEVER modified or
re-saved. Rerun this any time you update the tracker:

    python3 parse.py                     # auto-locates the workbook
    python3 parse.py /path/to/book.xlsm  # explicit path

It prints a per-sheet record count so you can check against the workbook's
own live counters.
"""

import datetime
import json
import os
import sys

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

WORKBOOK_NAME = "The_Merz_Legal_Crystal_Ball_.xlsm"

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(HERE, "docs", "data.json")

# Data sheets: headers on row 4, data from row 5. Second column is always the
# "entity" (Company / Plaintiff / Challenger); rows without it are year-divider
# rows or trailing blanks and are skipped.
DATA_SHEETS = [
    "DOJ Settlements",
    "OPDP Warning Letters",
    "OPDP Untitled Letters",
    "Direct Quotes",
    "Pharma v Pharma Lawsuits",
    "NAD Challenges",
    # Channel sheets
    "TV Commercials",
    "YouTube, Podcasts & Interviews",
    "Online Video",
    "Social Media",
    "Banner & Display Ads",
    "Sales Aids & Print",
    "Email",
    "Conferences & In-Person",
    "Investor & Press",
]

GLOSSARY_SHEET = "Glossary"

HEADER_ROW = 4
FIRST_DATA_ROW = 5


def locate_workbook(argv):
    """Return the workbook path, searching a few sensible locations."""
    if len(argv) > 1:
        p = argv[1]
        if not os.path.isfile(p):
            sys.exit(f"ERROR: workbook not found at {p!r}")
        return p
    candidates = [
        os.path.join(HERE, WORKBOOK_NAME),
        os.path.join(os.path.expanduser("~"), "Downloads", WORKBOOK_NAME),
        os.path.join(os.getcwd(), WORKBOOK_NAME),
    ]
    for c in candidates:
        if os.path.isfile(c):
            return c
    sys.exit(
        "ERROR: could not find the workbook. Pass its path explicitly:\n"
        "    python3 parse.py /path/to/%s" % WORKBOOK_NAME
    )


def clean(value):
    """Normalize a cell value for JSON output. Returns None for empties."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        # Dates in this workbook carry no meaningful time component.
        if value.hour or value.minute or value.second:
            return value.isoformat()
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return value


def read_headers(ws, ncols):
    """Read the header row, trimming trailing empty header cells."""
    row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                            max_col=ncols, values_only=True))
    headers = [clean(c) for c in row]
    # Trim trailing Nones so we don't carry phantom columns.
    while headers and headers[-1] is None:
        headers.pop()
    return headers


def parse_data_sheet(ws):
    """Yield one dict per real record in a data sheet."""
    headers = read_headers(ws, ws.max_column)
    ncols = len(headers)
    if ncols < 2:
        return
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=ncols,
                            values_only=True):
        # The second column is the entity (Company/Plaintiff/Challenger).
        # Empty there => year-divider row or trailing blank => skip.
        entity = clean(row[1]) if len(row) > 1 else None
        if entity is None:
            continue
        record = {"source_sheet": ws.title}
        for header, raw in zip(headers, row):
            if header is None:
                continue
            v = clean(raw)
            if v is not None:
                record[header] = v
        yield record


def parse_glossary(ws):
    """Yield glossary entries: Term, Category, Definition, Where It Appears.

    Column H (macro aliases) and anything past 'Where It Appears' is ignored.
    """
    # Fixed 4-column layout per the workbook spec (A-D).
    cols = ["Term", "Category", "Definition", "Where It Appears"]
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=4,
                            values_only=True):
        term = clean(row[0]) if row else None
        if term is None:
            continue
        entry = {}
        for name, raw in zip(cols, row):
            v = clean(raw)
            if v is not None:
                entry[name] = v
        yield entry


def main():
    path = locate_workbook(sys.argv)
    print(f"Reading (read-only): {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_titles = set(wb.sheetnames)

    records = []
    counts = {}
    for name in DATA_SHEETS:
        if name not in sheet_titles:
            print(f"  WARNING: expected sheet '{name}' not found; skipping")
            continue
        rows = list(parse_data_sheet(wb[name]))
        records.extend(rows)
        counts[name] = len(rows)

    glossary = []
    if GLOSSARY_SHEET in sheet_titles:
        glossary = list(parse_glossary(wb[GLOSSARY_SHEET]))

    wb.close()

    payload = {"records": records, "glossary": glossary}
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    # Report ------------------------------------------------------------------
    print("\nPer-sheet record counts (check against the workbook counters):")
    width = max(len(n) for n in counts) if counts else 0
    for name in DATA_SHEETS:
        if name in counts:
            print(f"  {name.ljust(width)}  {counts[name]:>4}")
    total = sum(counts.values())
    print(f"  {'-' * width}  ----")
    print(f"  {'TOTAL records'.ljust(width)}  {total:>4}")
    print(f"  {'Glossary terms'.ljust(width)}  {len(glossary):>4}")
    print(f"\nWrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
